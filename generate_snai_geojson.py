"""
Genera aree-snai-perimetri.geojson: poligoni dissolti per area SNAI.
Pipeline: legge ISTAT codes da Excel → scarica confini comunali ISTAT → dissolve per area → semplifica → esporta.
"""
import openpyxl
import geopandas as gpd
import pandas as pd
import json
import requests
import os
import sys

EXCEL_08 = r'C:\Users\aalbe\Desktop\Code\Aree Interne Italia\DATA\08_elenco-aree-comuni.xlsx'
EXCEL_72 = r'C:\Users\aalbe\Desktop\Code\Aree Interne Italia\DATA\elenco_aree_snai_14-20-e-21-27.xlsx'
COMUNI_CACHE = r'C:\Users\aalbe\Desktop\Code\Aree Interne Italia\DATA\comuni_italia.geojson'
OUT_GEOJSON = r'C:\Users\aalbe\Desktop\Code\Aree Interne Italia\aree-snai-perimetri.geojson'

# ── 1. Estrai comuni con codice ISTAT e nome area ─────────────────────────────

def read_comuni(wb, sheet, istat_col, area_col, status, skip=4):
    ws = wb[sheet]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < skip:
            continue
        istat = r[istat_col] if len(r) > istat_col else None
        area  = r[area_col]  if len(r) > area_col  else None
        if istat and area and str(istat).strip():
            rows.append({'istat': str(int(istat)).zfill(6), 'area': str(area).strip(), 'status': status})
    return rows

print("Lettura Excel 08_elenco-aree-comuni.xlsx …")
wb08 = openpyxl.load_workbook(EXCEL_08, read_only=True, data_only=True)
comuni = []
# SI finanziate: ISTAT=col2 (PRO_COM_T), area=col10 (Aree SNAI 2021-2027)
comuni += read_comuni(wb08, 'Comuni Aree 21-27 finanziate',    istat_col=2, area_col=10, status='SI')
# CONF confermate: stessa struttura
comuni += read_comuni(wb08, 'Comuni Aree 14-20 confermate',     istat_col=2, area_col=10, status='CONF')
# NO non finanziate: area=col9
comuni += read_comuni(wb08, 'Comuni 13 Aree 21-27 NO finanz.', istat_col=2, area_col=9,  status='NO')

print(f"  {len(comuni)} comuni estratti (SI/CONF/NO)")

# ── 2. Le 5 aree solo-14-20 dall'altro Excel ──────────────────────────────────
# Foglio "Comuni 72 Aree SNAI 14-20": col 0 = Codice Comune (6 cifre), col 8 = area SNAI
# Riga 0: vuota, riga 1: intestazioni, dati da riga 2.
SOLO_1420_AREE = {
    'Alta Valtellina', 'Val Bormida', 'Val di Lanzo', "Val d'Ossola", 'Valli Grana e Maira'
}
print("Lettura elenco_aree_snai_14-20-e-21-27.xlsx …")
wb72 = openpyxl.load_workbook(EXCEL_72, read_only=True, data_only=True)
rows72 = read_comuni(wb72, 'Comuni 72 Aree SNAI 14-20', istat_col=0, area_col=8, status='1420', skip=2)
rows72 = [r for r in rows72 if r['area'] in SOLO_1420_AREE]
print(f"  {len(rows72)} comuni delle 5 aree solo-14-20")
comuni += rows72

df = pd.DataFrame(comuni)
print(f"\nTotale comuni: {len(df)}")
print(df.groupby('status').agg(aree=('area','nunique'), comuni=('istat','count')))

# ── 3. Scarica confini comunali ISTAT (openpolis) ────────────────────────────
COMUNI_URL = 'https://raw.githubusercontent.com/openpolis/geojson-italy/master/geojson/limits_IT_municipalities.geojson'

if not os.path.exists(COMUNI_CACHE):
    print(f"\nDownload confini comunali (~30MB) → {COMUNI_CACHE} …")
    r = requests.get(COMUNI_URL, stream=True, timeout=120)
    r.raise_for_status()
    total = 0
    with open(COMUNI_CACHE, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024*256):
            f.write(chunk)
            total += len(chunk)
            print(f"  {total/1e6:.1f} MB scaricati…", end='\r')
    print(f"\n  Scaricati {total/1e6:.1f} MB")
else:
    print(f"\nConfini comunali già in cache: {COMUNI_CACHE}")

print("Caricamento GeoDataFrame …")
gdf = gpd.read_file(COMUNI_CACHE)
print(f"  {len(gdf)} comuni nel GeoJSON. Colonne: {list(gdf.columns[:8])}")

# ── 4. Identifica colonna ISTAT nel GeoJSON ───────────────────────────────────
# Priorità: com_istat_code (6 cifre) > com_istat_code_num (int) > fallback
istat_col_geo = None
for candidate in ['com_istat_code', 'com_istat_code_num', 'cod_reg_istat', 'istat']:
    if candidate in gdf.columns:
        istat_col_geo = candidate
        sample = str(gdf[istat_col_geo].iloc[0])
        print(f"  Colonna ISTAT: {istat_col_geo} (esempio: {sample})")
        break

if istat_col_geo is None:
    print("  WARN: colonna ISTAT non identificata automaticamente, uso prima colonna stringa 6-cifre")
    for col in gdf.columns:
        vals = gdf[col].dropna().astype(str)
        if vals.str.match(r'^\d{5,6}$').mean() > 0.8:
            istat_col_geo = col
            print(f"  Colonna ISTAT rilevata: {col}")
            break

if istat_col_geo is None:
    print("ERRORE: colonna ISTAT non trovata. Colonne disponibili:", list(gdf.columns))
    sys.exit(1)

# Normalizza codice ISTAT nel GeoJSON a 6 cifre
gdf['istat_norm'] = gdf[istat_col_geo].astype(str).str.strip().str.zfill(6)
df['istat'] = df['istat'].str.strip().str.zfill(6)

# ── 5. Join: assegna area SNAI ad ogni comune ─────────────────────────────────
merged = gdf.merge(df[['istat', 'area', 'status']], left_on='istat_norm', right_on='istat', how='inner')
print(f"\nComuni abbinati: {len(merged)} su {len(df)} cercati")
mancanti = len(df) - len(merged)
if mancanti > 0:
    matched_set = set(merged['istat'])
    miss = df[~df['istat'].isin(matched_set)]
    print(f"  Mancanti: {mancanti} — esempi: {miss['istat'].tolist()[:10]}")

# ── 6. Dissolvi per area ──────────────────────────────────────────────────────
print("\nDissolving poligoni per area SNAI …")
dissolved = merged.dissolve(by='area', as_index=False)[['area', 'status', 'geometry']]
# Aggiungi metadati dalle liste conosciute (pop, n_comuni)
area_meta = df.groupby('area').agg(n_comuni_istat=('istat','count'), status=('status','first')).reset_index()
dissolved = dissolved.merge(area_meta[['area','n_comuni_istat']], on='area', how='left')
print(f"  {len(dissolved)} aree dissolte")

# ── 7. Semplifica geometrie ───────────────────────────────────────────────────
print("Semplificazione geometrie (tolerance=0.002°) …")
dissolved = dissolved.to_crs('EPSG:4326')
dissolved['geometry'] = dissolved['geometry'].simplify(tolerance=0.002, preserve_topology=True)

# ── 8. Esporta GeoJSON aree dissolte ─────────────────────────────────────────
print(f"\nEsporto -> {OUT_GEOJSON}")
dissolved.to_file(OUT_GEOJSON, driver='GeoJSON')
size_kb = os.path.getsize(OUT_GEOJSON) / 1024
print(f"  Dimensione: {size_kb:.0f} KB")

# ── 9. Esporta GeoJSON comuni individuali ─────────────────────────────────────
OUT_COMUNI_GEOJSON = r'C:\Users\aalbe\Desktop\Code\Aree Interne Italia\comuni-snai-perimetri.geojson'
print(f"\nEsporto comuni individuali -> {OUT_COMUNI_GEOJSON}")
comuni_web = merged[['area', 'status', 'name', 'com_istat_code', 'geometry']].copy()
comuni_web = comuni_web.rename(columns={'name': 'comune', 'com_istat_code': 'procom'})
comuni_web = comuni_web.to_crs('EPSG:4326')
comuni_web['geometry'] = comuni_web['geometry'].simplify(tolerance=0.003, preserve_topology=True)
comuni_web.to_file(OUT_COMUNI_GEOJSON, driver='GeoJSON')
size_kb2 = os.path.getsize(OUT_COMUNI_GEOJSON) / 1024
print(f"  Dimensione: {size_kb2:.0f} KB")

print("\nCompletato.")
